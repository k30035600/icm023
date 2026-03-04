# -*- coding: utf-8 -*-
"""
ICM023 고객관리 - Flask 서버
information/insurance/accident/counsel xlsx CRUD 및 accident/counsel 추가 저장.
"""
import logging
import os
import re
import json
import shutil
from datetime import datetime, date, timezone, timedelta

KST = timezone(timedelta(hours=9))
SERVER_BOOT_TIME = datetime.now(KST).strftime('%Y-%m-%d %H:%M')


# 204/304 등 정상 요청 로그를 터미널에 찍지 않음 (에러만 출력)
logging.getLogger('werkzeug').setLevel(logging.ERROR)

import traceback
import threading
import io
import zipfile
from flask import Flask, request, jsonify, send_from_directory, send_file
import openpyxl

excel_lock = threading.Lock()

app = Flask(__name__, static_folder='.', static_url_path='')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_DIR = os.path.join(BASE_DIR, '.source')
os.makedirs(SOURCE_DIR, exist_ok=True)

INFORMATION_FILE = os.path.join(SOURCE_DIR, 'icm023_information.xlsx')
INSURANCE_FILE   = os.path.join(SOURCE_DIR, 'icm023_insurance.xlsx')
ACCIDENT_FILE    = os.path.join(SOURCE_DIR, 'icm023_accident.xlsx')
COUNSEL_FILE     = os.path.join(SOURCE_DIR, 'icm023_counsel.xlsx')

# ---------------------------------------------------------------------------
# 유틸리티
# ---------------------------------------------------------------------------

def normalize_jumin(v):
    if v is None:
        return ''
    s = re.sub(r'\D', '', str(v).strip())
    return s[:13] if s else ''

def str_to_yyyymmdd(s):
    """'yy/mm/dd' 또는 'yyyy/mm/dd', 'yyyy-mm' 문자열 → YYYYMMDD 정수 (실패 시 None)."""
    if not s or not isinstance(s, str):
        return None
    parts = re.sub(r'\D+', ' ', s.strip()).split()
    if len(parts) < 3:
        if len(parts) == 2:
            try:
                y, m = int(parts[0]), int(parts[1])
                if y < 100:
                    y += 2000 if y < 50 else 1900
                if 1 <= m <= 12:
                    return y * 10000 + m * 100 + 1
            except (ValueError, TypeError):
                pass
        return None
    try:
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        if y < 100:
            y += 2000 if y < 50 else 1900
        if 1 <= m <= 12 and 1 <= d <= 31:
            return y * 10000 + m * 100 + d
    except (ValueError, TypeError):
        pass
    return None

def yyyymmdd_to_excel_date(n):
    """YYYYMMDD 정수 → Excel 날짜 시리얼 (실패 시 None)."""
    if n is None or not isinstance(n, (int, float)):
        return None
    n = int(n)
    if n < 19000101 or n > 29991231:
        return None
    y, m, d = n // 10000, (n // 100) % 100, n % 100
    try:
        return (datetime(y, m, d) - datetime(1899, 12, 30)).days
    except ValueError:
        return None

def _date_value_for_sheet(h, v):
    """배정일자·종료일자·납입만기 문자열을 Excel 날짜 시리얼로 변환."""
    if v is None or v == '':
        return ''
    if h in ('배정일자', '종료일자', '납입만기') and isinstance(v, str):
        n = str_to_yyyymmdd(v)
        if n is not None:
            serial = yyyymmdd_to_excel_date(n)
            return serial if serial is not None else v
    return v

def _amount_to_number(v):
    """'1,234' 형태의 금액 문자열을 float으로 변환. 변환 불가 시 원값 반환."""
    if isinstance(v, str) and re.match(r'^[\d,]+$', v):
        try:
            return float(v.replace(',', ''))
        except ValueError:
            pass
    return v

def _load_or_create_wb(file_path, default_headers, sheet_title='Sheet1'):
    """xlsx 로드(헤더 포함) 또는 신규 워크북+헤더 행 생성. (wb, ws, headers) 반환."""
    if os.path.isfile(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title
        headers = list(default_headers)
        ws.append(headers)
    return wb, ws, headers

def _get_real_max_row(ws):
    real_max = 1
    empty_cnt = 0
    # max_row가 백만단위로 오동작할 때를 대비하여 비어있는 셀을 카운트
    for r in range(1, min(ws.max_row, 1048576) + 1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, 6)):
            real_max = r
            empty_cnt = 0
        else:
            empty_cnt += 1
            if empty_cnt > 50:
                break
    return real_max

def _append_to_ws(ws, row_values):
    real_max = _get_real_max_row(ws)
    for c, val in enumerate(row_values, 1):
        ws.cell(row=real_max + 1, column=c, value=val)

def _delete_rows_by_key(ws, headers, key_name, key_jumin):
    """ws에서 계약자명+주민번호가 일치하는 행을 모두 삭제. 삭제 건수 반환."""
    name_col = next((i for i, h in enumerate(headers) if h == '계약자명'), None)
    jumin_col = next((i for i, h in enumerate(headers) if h in ('주민등록번호', '주민번호')), None)
    if name_col is None or jumin_col is None:
        return 0
    to_delete = [
        r for r in range(2, _get_real_max_row(ws) + 1)
        if str(ws.cell(row=r, column=name_col + 1).value or '').strip() == key_name
        and normalize_jumin(ws.cell(row=r, column=jumin_col + 1).value) == key_jumin
    ]
    for r in reversed(to_delete):
        ws.delete_rows(r, 1)
    return len(to_delete)

def _build_log_row(headers, row_dict):
    """accident/counsel 행 값 배열 생성 (기록일자·기록시간 변환 포함)."""
    row_values = []
    for h in headers:
        v = row_dict.get(h)
        if h == '기록일자' and v is not None:
            serial = yyyymmdd_to_excel_date(v)
            row_values.append(serial if serial is not None else v)
        elif h == '기록시간' and v is not None and isinstance(v, (int, float)):
            row_values.append(float(v))
        else:
            row_values.append(v if v is not None else '')
    return row_values

# ---------------------------------------------------------------------------
# 라우트
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/api/status')
def api_status():
    return jsonify({'boot_time': SERVER_BOOT_TIME})

@app.route('/favicon.ico')
def favicon():
    """favicon 요청 시 파일 없으면 204 반환하여 404 로그 방지."""
    path = os.path.join(BASE_DIR, 'favicon.ico')
    if os.path.isfile(path):
        return send_from_directory(BASE_DIR, 'favicon.ico')
    return '', 204

@app.route('/<path:path>')
def static_file(path):
    # .source 폴더 직접 접근 차단 (개인정보 보호)
    if path.startswith('.source') or '/.source/' in path:
        return '', 403
    return send_from_directory(BASE_DIR, path)

# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------

@app.route('/api/backup_restore', methods=['POST'])
def api_backup_restore():
    """
    백업/복원 기능:
    backup: 현재 xlsx 파일들을 .source 폴더의 json 파일로 저장 (백업)
    restore: .source 폴더의 json 파일들로 현재 xlsx 파일을 덮어쓰기 (복원)
    """
    if request.mimetype == 'multipart/form-data':
        pwd = request.form.get('password', '')
        action = request.form.get('action')
        file = request.files.get('file')
    else:
        req = request.get_json(force=True) or {}
        pwd = req.get('password', '')
        action = req.get('action')
        file = None

    if action == 'backup' and pwd != 'backup':
        return jsonify({"error": "패스워드가 일치하지 않습니다."}), 403
    elif action == 'restore' and pwd != 'restore':
        return jsonify({"error": "패스워드가 일치하지 않습니다."}), 403

    files_map = {
        'icm023_products': os.path.join(SOURCE_DIR, 'icm023_products.xlsx'),
        'icm023_information': INFORMATION_FILE,
        'icm023_insurance': INSURANCE_FILE,
        'icm023_accident': ACCIDENT_FILE,
        'icm023_counsel': COUNSEL_FILE
    }

    try:
        with excel_lock:
            if action == 'backup':
                memory_file = io.BytesIO()
                with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for name, xlsx_path in files_map.items():
                        if os.path.isfile(xlsx_path):
                            zf.write(xlsx_path, os.path.basename(xlsx_path))
                memory_file.seek(0)
                return send_file(
                    memory_file,
                    download_name='backup.zip',
                    as_attachment=True,
                    mimetype='application/zip'
                )

            elif action == 'restore':
                if not file or not file.filename.endswith('.zip'):
                    return jsonify({"error": "ZIP 백업 파일이 제공되지 않았거나 형식이 잘못되었습니다."}), 400
                
                try:
                    with zipfile.ZipFile(file, 'r') as zf:
                        for info in zf.infolist():
                            fname = os.path.basename(info.filename)
                            for name, xlsx_path in files_map.items():
                                if fname == os.path.basename(xlsx_path):
                                    with open(xlsx_path, 'wb') as f:
                                        f.write(zf.read(info.filename))
                                    break
                    return jsonify({"ok": True, "message": "압축 파일 복원이 성공적으로 완료되었습니다."})
                except zipfile.BadZipFile:
                    return jsonify({"error": "유효한 ZIP 압축 파일이 아닙니다."}), 400
            else:
                return jsonify({"error": "알 수 없는 작업입니다."}), 400
    except Exception as e:
        logging.exception(f"Error during {action}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/data/<name>', methods=['GET'])
def api_get_data(name):
    """
    엑셀 파일을 읽어 JSON으로 단번에 반환하는 통합 API (프론트엔드 최적화)
    name: 'information', 'insurance', 'accident', 'counsel', 'products'
    """
    file_map = {
        'information': INFORMATION_FILE,
        'insurance': INSURANCE_FILE,
        'accident': ACCIDENT_FILE,
        'counsel': COUNSEL_FILE,
        'products': os.path.join(SOURCE_DIR, 'icm023_products.xlsx')
    }
    file_path = file_map.get(name)
    if not file_path:
        return jsonify({"error": "Unknown data name"}), 400

    if not os.path.isfile(file_path):
        return jsonify({"headers": [], "rows": []})

    try:
        with excel_lock:
            # read_only=True: 스트리밍 방식으로 읽어 비정상 max_row 문제 우회
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if not wb.sheetnames:
                wb.close()
                return jsonify({"headers": [], "rows": []})
            ws = wb[wb.sheetnames[0]]

            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
                # 헤더 정규화 (엑셀과 프론트엔드의 차이 보정)
                headers = []
                for cell in header_row:
                    h = str(cell).strip() if cell is not None else ''
                    if h == '주민등록번호': h = '주민번호'
                    elif h == '보험종류': h = '상품종류'
                    elif h == 'Q/A': h = '문/답'
                    headers.append(h)
            except StopIteration:
                wb.close()
                return jsonify({"headers": [], "rows": []})

            data_rows = []
            file_row_index = 0
            empty_streak = 0
            for row in rows_iter:
                row_dict = {}
                is_empty = True
                for i, v in enumerate(row):
                    if i < len(headers):
                        val = v
                        if isinstance(val, datetime):
                            val = val.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(val, date):
                            val = val.strftime('%Y-%m-%d')
                        elif isinstance(val, (int, float)):
                            # 날짜 형식인 일련번호인지 등은 프론트에서 처리하므로 원본을 넘김
                            pass

                        if val is not None and str(val).strip() != '':
                            is_empty = False
                        row_dict[headers[i]] = val

                if not is_empty:
                    # accident 등 데이터 조작에 필요한 엑셀 실제 열 번호 (0-based)
                    row_dict['_fileRowIndex'] = file_row_index
                    data_rows.append(row_dict)
                    empty_streak = 0
                else:
                    empty_streak += 1
                    if empty_streak > 50:
                        break  # 연속 빈 행 50개 초과 시 조기 종료

                file_row_index += 1

            wb.close()
            return jsonify({
                "headers": headers,
                # 중복 필터링 제거 (이미 append 단계에서 걸렀으므로)
                "rows": data_rows
            })
    except Exception as e:
        logging.exception(f"Error reading {name} data")
        return jsonify({"error": str(e)}), 500

# ---------------------------------------------------------------------------

@app.route('/api/information', methods=['POST'])
def api_information():
    """information xlsx: action=insert|update|delete."""
    try:
        data = request.get_json(force=True) or {}
        action = (data.get('action') or '').strip()
        if action not in ('insert', 'update', 'delete'):
            return jsonify({'ok': False, 'error': 'action은 insert/update/delete 중 하나여야 합니다.'}), 400

        default_hdrs = data.get('headers') or ['계약자명', '주민번호', '배정일자', '종료일자', '문자전송', 'DB종류', '연락처']
        with excel_lock:
            wb, ws, headers = _load_or_create_wb(INFORMATION_FILE, default_hdrs)
            row_index = data.get('rowIndex')
            row_dict  = data.get('row') or {}

            ky_name = str(row_dict.get('계약자명') or '').strip()
            pi_name = str(row_dict.get('피보험자명') or '').strip()
            if ky_name and ky_name == pi_name:
                row_dict['피주민번호'] = ''
                row_dict['피연락처'] = ''
                row_dict['피직업'] = ''

            # 중복 체크 (계약자명 + 주민번호 + 피보험자명)
            if action in ('insert', 'update'):
                name_idx = next((i for i, h in enumerate(headers) if h == '계약자명'), None)
                jumin_idx = next((i for i, h in enumerate(headers) if h == '주민번호'), None)
                pi_idx = next((i for i, h in enumerate(headers) if h == '피보험자명'), None)
                if name_idx is not None and jumin_idx is not None and pi_idx is not None:
                    check_name = ky_name
                    check_jumin = normalize_jumin(row_dict.get('주민번호'))
                    check_pi = pi_name
                    target_excel_row = (2 + int(row_index)) if (action == 'update' and row_index is not None) else -1
                    
                    for r in range(2, _get_real_max_row(ws) + 1):
                        if r == target_excel_row: continue
                        ex_n = str(ws.cell(row=r, column=name_idx + 1).value or '').strip()
                        ex_j = normalize_jumin(ws.cell(row=r, column=jumin_idx + 1).value)
                        ex_p = str(ws.cell(row=r, column=pi_idx + 1).value or '').strip()
                        if ex_n == check_name and ex_j == check_jumin and ex_p == check_pi:
                            msg = f'이미 등록된 계약입니다. ({check_name} / {row_dict.get("주민번호")} / 피:{check_pi})'
                            return jsonify({'ok': False, 'error': msg}), 400

            if action == 'delete':
                if row_index is None:
                    return jsonify({'ok': False, 'error': '삭제 시 rowIndex 필요'}), 400
                excel_row = 2 + int(row_index)
                if excel_row < 2 or excel_row > _get_real_max_row(ws):
                    return jsonify({'ok': False, 'error': '유효하지 않은 행'}), 400
                ws.delete_rows(excel_row, 1)
            elif action == 'update':
                if row_index is None:
                    return jsonify({'ok': False, 'error': '수정 시 rowIndex 필요'}), 400
                excel_row = 2 + int(row_index)
                if excel_row < 2 or excel_row > _get_real_max_row(ws):
                    return jsonify({'ok': False, 'error': '유효하지 않은 행'}), 400
                for col, h in enumerate(headers, 1):
                    ws.cell(row=excel_row, column=col, value=_date_value_for_sheet(h, row_dict.get(h, '')))
            else:
                _append_to_ws(ws, [_date_value_for_sheet(h, row_dict.get(h, '')) for h in headers])

            wb.save(INFORMATION_FILE)
            wb.close()
        return jsonify({'ok': True, 'action': action})
    except Exception as e:
        logging.exception("Error in /api/information")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/insurance', methods=['POST'])
def api_insurance():
    """insurance xlsx: action=insert|update|delete. update/delete는 key로 행 식별."""
    try:
        data = request.get_json(force=True) or {}
        action = (data.get('action') or '').strip()
        if action not in ('insert', 'update', 'delete'):
            return jsonify({'ok': False, 'error': 'action은 insert/update/delete 중 하나여야 합니다.'}), 400

        key_raw   = data.get('key') or {}
        key_name  = (key_raw.get('계약자명') or '').strip()
        key_jumin = normalize_jumin(key_raw.get('주민번호') or key_raw.get('주민등록번호'))

        default_hdrs = ['계약자명', '주민번호', '상품종류', '상품명', '증권번호', '납입만기', '납입보험료', '총납입보험료', '해지환급금', '대출금']
        with excel_lock:
            wb, ws, headers = _load_or_create_wb(INSURANCE_FILE, default_hdrs)
            if '보험종류' in headers and '상품종류' not in headers:
                headers = ['상품종류' if h == '보험종류' else h for h in headers]

            row_dict = data.get('row') or {}

            def _make_row(src):
                return [_amount_to_number(_date_value_for_sheet(h, src.get(h, '') or (src.get('상품종류', '') if h == '보험종류' else ''))) for h in headers]

            def row_matches(r):
                return str(r.get('계약자명') or '').strip() == key_name and \
                       normalize_jumin(r.get('주민번호') or r.get('주민등록번호')) == key_jumin

            # 중복 체크 (계약자명 + 주민번호 + 증권번호)
            if action in ('insert', 'update'):
                 target_file_row = int(data.get('fileRowIndex')) if (action == 'update' and 'fileRowIndex' in data) else -1
                 check_no = str(row_dict.get('증권번호') or '').strip()
                 check_n = str(row_dict.get('계약자명') or '').strip()
                 check_j = normalize_jumin(row_dict.get('주민번호'))
                 
                 n_idx = next((i for i, h in enumerate(headers) if h == '계약자명'), None)
                 j_idx = next((i for i, h in enumerate(headers) if h == '주민번호'), None)
                 no_idx = next((i for i, h in enumerate(headers) if h == '증권번호'), None)
                 
                 if n_idx is not None and j_idx is not None and no_idx is not None:
                     for r in range(2, _get_real_max_row(ws) + 1):
                         if (r - 2) == target_file_row: continue
                         ex_n = str(ws.cell(row=r, column=n_idx + 1).value or '').strip()
                         ex_j = normalize_jumin(ws.cell(row=r, column=j_idx + 1).value)
                         ex_no = str(ws.cell(row=r, column=no_idx + 1).value or '').strip()
                         if ex_n == check_n and ex_j == check_j and ex_no == check_no:
                             return jsonify({'ok': False, 'error': f'이미 등록된 증권번호입니다. ({check_no})'}), 400

            if action == 'delete':
                if 'fileRowIndex' in data:
                    ws.delete_rows(2 + int(data['fileRowIndex']), 1)
                else:
                    _delete_rows_by_key(ws, headers, key_name, key_jumin)
            elif action == 'update':
                updated = False
                for r in range(2, _get_real_max_row(ws) + 1):
                    row_obj = {h: ws.cell(row=r, column=c).value for c, h in enumerate(headers, 1)}
                    if row_matches(row_obj):
                        merged = {h: row_dict.get(h, row_obj.get(h, '')) for h in headers}
                        for c, h in enumerate(headers, 1):
                            ws.cell(row=r, column=c, value=_amount_to_number(_date_value_for_sheet(h, merged[h])))
                        updated = True
                        break
                if not updated:
                    _append_to_ws(ws, _make_row(row_dict))
            else:
                _append_to_ws(ws, _make_row(row_dict))

            wb.save(INSURANCE_FILE)
            wb.close()
        return jsonify({'ok': True, 'action': action})
    except Exception as e:
        logging.exception("Error in /api/insurance")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/accident', methods=['POST'])
def accident_append():
    """accident xlsx: action=deleteRow(행 인덱스) | delete(키) | append(기본)."""
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'ok': False, 'error': 'JSON 없음'}), 400
        action = (data.get('action') or '').strip()

        if action == 'deleteRow':
            row_index = data.get('rowIndex')
            if row_index is None:
                return jsonify({'ok': False, 'error': '삭제할 행의 rowIndex가 필요합니다.'}), 400
            if not os.path.isfile(ACCIDENT_FILE):
                return jsonify({'ok': True, 'deleted': 0})
            with excel_lock:
                wb = openpyxl.load_workbook(ACCIDENT_FILE)
                ws = wb.active
                excel_row = 2 + int(row_index)
                if excel_row < 2 or excel_row > _get_real_max_row(ws):
                    wb.close()
                    return jsonify({'ok': False, 'error': '유효하지 않은 행 인덱스입니다.'}), 400
                ws.delete_rows(excel_row, 1)
                wb.save(ACCIDENT_FILE)
                wb.close()
            return jsonify({'ok': True, 'deleted': 1})
            
        if action == 'updateRow':
            row_index = data.get('rowIndex')
            if row_index is None:
                return jsonify({'ok': False, 'error': '수정할 행의 rowIndex가 필요합니다.'}), 400
            row_dict = data.get('row') or {}
            if not os.path.isfile(ACCIDENT_FILE):
                return jsonify({'ok': False, 'error': '파일이 없습니다.'}), 400
            
            with excel_lock:
                wb = openpyxl.load_workbook(ACCIDENT_FILE)
                ws = wb.active
                excel_row = 2 + int(row_index)
                if excel_row < 2 or excel_row > _get_real_max_row(ws):
                    wb.close()
                    return jsonify({'ok': False, 'error': '유효하지 않은 행 인덱스입니다.'}), 400
                
                hdrs = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
                for col, h in enumerate(hdrs, 1):
                    if h in row_dict:
                        ws.cell(row=excel_row, column=col, value=row_dict[h])
                
                wb.save(ACCIDENT_FILE)
                wb.close()
            return jsonify({'ok': True, 'updated': 1})

        if action == 'delete':
            key_raw   = data.get('key') or {}
            key_name  = (key_raw.get('계약자명') or '').strip()
            key_jumin = normalize_jumin(key_raw.get('주민번호') or key_raw.get('주민등록번호'))
            if not os.path.isfile(ACCIDENT_FILE):
                return jsonify({'ok': True, 'deleted': 0})
            with excel_lock:
                wb = openpyxl.load_workbook(ACCIDENT_FILE)
                ws = wb.active
                headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
                deleted = _delete_rows_by_key(ws, headers, key_name, key_jumin)
                wb.save(ACCIDENT_FILE)
                wb.close()
            return jsonify({'ok': True, 'deleted': deleted})

        default_hdrs  = ['계약자명', '주민번호', '구분', '입력일자', '기록일자', '기록시간', '문/답', '세부내용']
        headers       = data.get('headers') or default_hdrs
        rows_to_append = data.get('rowsToAppend') or []
        if not rows_to_append:
            return jsonify({'ok': True, 'appended': 0})

        today = date.today()
        today_serial = yyyymmdd_to_excel_date(
            today.year * 10000 + today.month * 100 + today.day
        )
        with excel_lock:
            wb, ws, _ = _load_or_create_wb(ACCIDENT_FILE, headers, sheet_title='accident')
            for row_dict in rows_to_append:
                row_dict = dict(row_dict)
                if '입력일자' in headers and (row_dict.get('입력일자') is None or row_dict.get('입력일자') == ''):
                    row_dict['입력일자'] = today_serial
                _append_to_ws(ws, _build_log_row(headers, row_dict))
            wb.save(ACCIDENT_FILE)
            wb.close()
        return jsonify({'ok': True, 'appended': len(rows_to_append)})
    except Exception as e:
        logging.exception("Error in /api/accident")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/counsel', methods=['POST'])
def counsel_append():
    """counsel xlsx: action=delete(키) | append(기본)."""
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'ok': False, 'error': 'JSON 없음'}), 400
        action = (data.get('action') or '').strip()

        if action == 'delete':
            key_raw   = data.get('key') or {}
            key_name  = (key_raw.get('계약자명') or '').strip()
            key_jumin = normalize_jumin(key_raw.get('주민번호') or key_raw.get('주민등록번호'))
            if not os.path.isfile(COUNSEL_FILE):
                return jsonify({'ok': True, 'deleted': 0})
            with excel_lock:
                wb = openpyxl.load_workbook(COUNSEL_FILE)
                ws = wb.active
                headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
                deleted = _delete_rows_by_key(ws, headers, key_name, key_jumin)
                wb.save(COUNSEL_FILE)
                wb.close()
            return jsonify({'ok': True, 'deleted': deleted})

        # append
        default_hdrs   = ['계약자명', '주민번호', '기록일자', '기록시간', '문/답', '세부내용']
        headers        = data.get('headers') or default_hdrs
        rows_to_append = data.get('rowsToAppend') or []
        if not rows_to_append:
            return jsonify({'ok': True, 'appended': 0})

        with excel_lock:
            wb, ws, _ = _load_or_create_wb(COUNSEL_FILE, headers, sheet_title='counsel')
            for row_dict in rows_to_append:
                _append_to_ws(ws, _build_log_row(headers, row_dict))
            wb.save(COUNSEL_FILE)
            wb.close()
        return jsonify({'ok': True, 'appended': len(rows_to_append)})
    except Exception as e:
        logging.exception("Error in /api/counsel")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/shutdown', methods=['POST'])
def shutdown():
    os._exit(0)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
