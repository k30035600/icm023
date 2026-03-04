import openpyxl
import os

files = ['icm023_information.xlsx', 'icm023_insurance.xlsx', 'icm023_accident.xlsx', 'icm023_counsel.xlsx']

for f in files:
    path = os.path.join('.source', f)
    if not os.path.exists(path): continue
    print(f"Loading {f}... ")
    
    # Load with read_only=False inside a try-except to handle max boundary issues
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        
        real_max = 1
        cnt_empty = 0
        for r in range(1, ws.max_row + 1):
            has_val = False
            for c in range(1, 10):
                cell_val = None
                try:
                    cell_val = ws.cell(row=r, column=c).value
                except Exception:
                    pass
                if cell_val is not None and str(cell_val).strip() != '':
                    has_val = True
                    break
            if has_val:
                real_max = r
                cnt_empty = 0
            else:
                cnt_empty += 1
                if cnt_empty > 50:
                    break
                    
        print(f"{f}: max_row={ws.max_row}, real_max={real_max}")
        if ws.max_row > real_max:
            amount = ws.max_row - real_max
            print(f"Deleting {amount} clean_rows")
            ws.delete_rows(real_max + 1, amount)
            wb.save(path)
            print("Saved.")
        wb.close()
    except Exception as e:
        print(f"Error on {f}: {e}")
