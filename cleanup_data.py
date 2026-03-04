import openpyxl
import os
import re
from datetime import datetime, date

SOURCE_DIR = '.source'

def excel_serial_to_yy_mm_dd(serial):
    try:
        # Excel's 1899-12-30 epoch
        d = date.fromordinal(date(1899, 12, 30).toordinal() + int(serial))
        return d.strftime('%y/%m/%d')
    except Exception:
        return str(serial)

def cleanup():
    files = {
        'icm023_information.xlsx': [],
        'icm023_insurance.xlsx': ['납입보험료', '총납입보험료', '해지환급금', '대출금'],
        'icm023_accident.xlsx': [],
        'icm023_counsel.xlsx': []
    }
    
    # 납입만기 column showing as numbers needs conversion to yy/mm/dd strings
    date_cols = ('납입만기',) 

    for fname, amt_cols in files.items():
        path = os.path.join(SOURCE_DIR, fname)
        if not os.path.exists(path): continue
        print(f"Cleaning {fname}...")
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            
            changed = False
            for r in range(2, ws.max_row + 1):
                for c, h in enumerate(headers, 1):
                    val = ws.cell(row=r, column=c).value
                    
                    # Fix amounts: empty/None -> 0
                    if h in amt_cols:
                        if val is None or str(val).strip() == '':
                            ws.cell(row=r, column=c, value=0)
                            changed = True
                        elif isinstance(val, str):
                            # Try converting string numbers with commas
                            clean_v = val.replace(',', '').strip()
                            if re.match(r'^-?\d+(\.\d+)?$', clean_v):
                                try:
                                    ws.cell(row=r, column=c, value=float(clean_v))
                                    changed = True
                                except: pass

                    # Fix 납입만기 serial numbers -> strings
                    if h in date_cols:
                        if isinstance(val, (int, float)) and val > 100: 
                            new_val = excel_serial_to_yy_mm_dd(val)
                            print(f"  Row {r} {h}: {val} -> {new_val}")
                            ws.cell(row=r, column=c, value=new_val)
                            changed = True

            if changed:
                wb.save(path)
            wb.close()
        except Exception as e:
            print(f"Error cleaning {fname}: {e}")
            
    print("Cleanup data consistency done.")

if __name__ == "__main__":
    cleanup()
